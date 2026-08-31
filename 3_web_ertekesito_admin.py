import streamlit as st
import pandas as pd
import firebase_admin
from firebase_admin import credentials, firestore
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
from collections import defaultdict

# --- KONFIGURÁCIÓ ---
ADMIN_JELSZO = "admin123"
st.set_page_config(page_title="Balettcipő Raktár", layout="wide")

# --- FIREBASE INDÍTÁSA ---
@st.cache_resource
def get_db():
    if not firebase_admin._apps:
        secrets = st.secrets["firestore"]
        cred_dict = dict(secrets)
        cred_dict["private_key"] = cred_dict["private_key"].replace("\\n", "\n")
        cred = credentials.Certificate(cred_dict)
        firebase_admin.initialize_app(cred)
    return firestore.client()

db = get_db()

# --- SEGÉDFÜGGVÉNYEK ---
def get_firebase_data():
    try:
        docs = db.collection("keszlet").stream()
        data = {}
        for doc in docs:
            d = doc.to_dict()
            menny = d.get("mennyiseg", 0)
            min_e = d.get("min_ertek", 0)
            try:
                menny = int(menny)
            except (ValueError, TypeError):
                menny = 0
            try:
                min_e = int(min_e)
            except (ValueError, TypeError):
                min_e = 0
                
            data[str(doc.id)] = {
                "mennyiseg": menny,
                "min_ertek": min_e
            }
        return data
    except Exception as e:
        st.error(f"Adatbázis hiba: {e}")
        return {}

def get_matrix(adatok, w):
    sizes = [str(i) for i in range(5, 15)]
    hardnesses = ["LGH", "SFT", "FLX", "SUP", "REG", "FRM", "STR", "XFR", "XST"]
    
    # Készítünk egy üres Mátrixot int típusokkal
    matrix = pd.DataFrame(0, index=hardnesses, columns=sizes)
    
    for m in sizes:
        for k in hardnesses:
            sku = f"{m}_{w}_{k}"
            termek_info = adatok.get(sku, {"mennyiseg": 0})
            if isinstance(termek_info, dict):
                val = termek_info.get("mennyiseg", 0)
            else:
                val = termek_info
            matrix.at[k, m] = int(val) if val is not None else 0
    
    # 1. Összesen sor
    matrix.loc["ÖSSZESEN"] = matrix.sum(axis=0)
    
    # 2. Reset index az első oszlop megnevezéséhez
    df = matrix.reset_index()
    df.rename(columns={"index": "Keménység"}, inplace=True)
    
    # 3. Összesen oszlop a sorok végére
    sizes_cols = [c for c in df.columns if c != "Keménység"]
    df["ÖSSZESEN"] = df[sizes_cols].sum(axis=1)
    
    return df

def szinezo(row):
    szinek = {
        "LGH": "#FFD1DC", "SFT": "#FFFFFF", "FLX": "#FF91A4", 
        "SUP": "#E0E0E0", "REG": "#FFFF00", "FRM": "#CD7F32", 
        "STR": "#00BFFF", "XFR": "#A6A6A6", "XST": "#FF4500" 
    }
    cell_value = str(row.iloc[0])
    
    if cell_value == "ÖSSZESEN": 
        return ['background-color: #f0f0f0; font-weight: bold'] * len(row)
    
    color = szinek.get(cell_value, "#FFFFFF")
    return [f'background-color: {color}; font-weight: bold'] * len(row)

def szinezo_admin(row, adatok, w):
    style = [''] * len(row)
    kem = str(row.iloc[0])
    
    if kem == "ÖSSZESEN": 
        return ['background-color: #f0f0f0; font-weight: bold'] * len(row)
    
    for i in range(1, len(row)):
        col_name = str(row.index[i])
        if col_name == "ÖSSZESEN":
            style[i] = 'background-color: #f0f0f0; font-weight: bold'
            continue
            
        sku = f"{col_name}_{w}_{kem}"
        info = adatok.get(sku, {"mennyiseg": 0, "min_ertek": 0})
        
        menny = info.get("mennyiseg", 0) if isinstance(info, dict) else info
        min_e = info.get("min_ertek", 0) if isinstance(info, dict) else 0
        
        if menny < min_e and min_e > 0:
            style[i] = 'background-color: #FF6666; color: white; font-weight: bold'
            
    return style

# --- RIPORT GENERÁLÁS ---
def generate_weekly_report(year, week):
    jan4 = datetime(year, 1, 4)
    start_date = jan4 + timedelta(days=(week - 1) * 7 - jan4.weekday())
    end_date = start_date + timedelta(days=6)
    
    naplo_docs = db.collection("naplo") \
        .where("datum", ">=", start_date.strftime("%Y-%m-%d")) \
        .where("datum", "<=", end_date.strftime("%Y-%m-%d")) \
        .stream()
    
    osszesites = defaultdict(int)
    for d in naplo_docs:
        doc = d.to_dict()
        key = (doc.get("datum"), doc.get("sku"), doc.get("tipus"))
        osszesites[key] += doc.get("darabszam", 0)

    wb = openpyxl.load_workbook("template.xlsx")
    ws = wb.active
    ws['O1'] = week
    
    for (datum, sku, tipus), mennyiseg in osszesites.items():
        datum_obj = datetime.strptime(datum, "%Y-%m-%d")
        nap_index = datum_obj.weekday() 
        col_offset = nap_index * 3 + 1
        
        start_row = 4 if tipus == "kiszedes" else 36 
        
        for r in range(start_row, start_row + 30):
            if ws.cell(row=r, column=col_offset).value is None:
                sku_parts = sku.split("_")
                ws.cell(row=r, column=col_offset, value=f"{sku_parts[0]}{sku_parts[1]}")
                ws.cell(row=r, column=col_offset+1, value=sku_parts[2])
                ws.cell(row=r, column=col_offset+2, value=mennyiseg)
                break
    
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

# --- APP LOGIKA ---
funkcio = st.sidebar.radio("Válassz felületet:", ["📱 Raktári Kiszedés", "📊 Értékesítő", "🔐 Admin"], key="nav")

if funkcio == "📱 Raktári Kiszedés":
    st.title("📱 Raktári Mozgás")
    
    adatok = get_firebase_data() 
    
    c1, c2, c3 = st.columns(3)
    meret = c1.selectbox("Méret:", [str(i) for i in range(5, 15)])
    szelesseg = c2.selectbox("Szélesség:", ["M", "W", "XW", "XXW"])
    kemenyseg = c3.selectbox("Keménység:", ["LGH", "SFT", "FLX", "SUP", "REG", "FRM", "STR", "XFR", "XST"])
    
    sku = f"{meret}_{szelesseg}_{kemenyseg}"
    
    akt_adat = adatok.get(sku, {"mennyiseg": 0})
    akt_mennyiseg = akt_adat.get("mennyiseg", 0) if isinstance(akt_adat, dict) else akt_adat
    
    st.write(f"Jelenlegi készlet: **{akt_mennyiseg}**")
    
    col1, col2 = st.columns(2)
    
    if col1.button("❌ Kiszedés"):
        db.collection("keszlet").document(sku).set({"mennyiseg": akt_mennyiseg - 1}, merge=True)
        db.collection("naplo").add({
            "datum": datetime.now().strftime("%Y-%m-%d"), 
            "sku": sku, 
            "tipus": "kiszedes", 
            "darabszam": 1
        })
        st.rerun()
        
    if col2.button("✅ Visszarakás"):
        db.collection("keszlet").document(sku).set({"mennyiseg": akt_mennyiseg + 1}, merge=True)
        db.collection("naplo").add({
            "datum": datetime.now().strftime("%Y-%m-%d"), 
            "sku": sku, 
            "tipus": "visszarakas", 
            "darabszam": 1
        })
        st.rerun()

    st.divider()
    st.subheader("📥 Heti riport export")
    ev, het = st.columns(2)
    ev_in = ev.number_input("Év", value=datetime.now().year)
    het_in = het.number_input("Hét", value=datetime.now().isocalendar()[1])
    if st.button("Riport készítése"):
        st.download_button("📥 Letöltés (Excel)", generate_weekly_report(ev_in, het_in), f"heti_riport_{ev_in}_W{het_in}.xlsx")

elif funkcio == "📊 Értékesítő":
    st.title("📊 Értékesítői Nézet")
    st.subheader("⚠️ ÉRTÉKESÍTHETŐ SPECIÁLIS KÉSZLET")
    col1, col2, col3 = st.columns(3)
    spec_data = {
        "V-LV": [["7W FLX", "5 pár"], ["6XXW REG", "1 pár"], ["8XW XTR", "1 pár"], ["11XW SUP", "1 pár"]],
        "U-LV": [["8W XFR", "1 pár"], ["8W REG", "2 pár"]],
        "U-DV": [["8M SFT", "8 pár"], ["8M STR", "1 pár"], ["9M STR", "3 pár"], ["9W STR", "2 pár"], ["8W XST", "1 pár"], ["11XXW XST", "1 pár"], ["11W FLX", "1 pár"]],
        "V-DV": [["4W SUP", "1 pár"], ["8W 1/2 XTR", "1 pár"], ["9XW 1/2 XTR", "2 pár"], ["10XW 1/2 XTR", "1 pár"], ["9XXW 2/3 REG", "1 pár"], ["9W REG H-CR", "1 pár"]]
    }
    with col1:
        st.info("### V-LV"); st.table(spec_data["V-LV"])
        st.info("### U-LV"); st.table(spec_data["U-LV"])
    with col2:
        st.success("### U-DV"); st.table(spec_data["U-DV"])
    with col3:
        st.success("### V-DV"); st.table(spec_data["V-DV"])
    
    st.divider()
    adatok = get_firebase_data()
    
    if st.button("📥 Összes leltár exportálása"):
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
            workbook = writer.book
            border_fmt = workbook.add_format({'border': 1})
            bold_border_fmt = workbook.add_format({'border': 1, 'bold': True})
            
            row_offset = 0
            for w in ["M", "W", "XW", "XXW"]:
                df = get_matrix(adatok, w)
                df.to_excel(writer, sheet_name="Keszlet", startrow=row_offset, index=False, header=True)
                worksheet = writer.sheets["Keszlet"]
                
                for r in range(df.shape[0] + 1):
                    for c in range(df.shape[1]):
                        if r == 0:
                            val = df.columns[c]
                            worksheet.write(row_offset + r, c, val, bold_border_fmt)
                        else:
                            val = df.iloc[r-1, c]
                            if r == df.shape[0]: 
                                worksheet.write(row_offset + r, c, val, bold_border_fmt)
                            else:
                                worksheet.write(row_offset + r, c, val, border_fmt)
                                
                row_offset += df.shape[0] + 2
                
        st.download_button("✅ Letöltés (Excel)", buffer.getvalue(), "Leltar_Osszes.xlsx")
    
    st.divider()
    for w in ["M", "W", "XW", "XXW"]:
        st.subheader(f"📦 {w} szélesség")
        df = get_matrix(adatok, w)
        st.dataframe(df.style.apply(szinezo, axis=1), use_container_width=True, hide_index=True)

elif funkcio == "🔐 Admin":
    st.title("🔐 Adminisztráció")
    if st.sidebar.text_input("Jelszó:", type="password") == ADMIN_JELSZO:
        adatok = get_firebase_data()
        
        for w in ["M", "W", "XW", "XXW"]:
            with st.expander(f"📦 {w} szélesség", expanded=True):
                df = get_matrix(adatok, w)
                
                # Admin nézetben egyetlen st.data_editor felületet használunk színezéssel
                edited_df = st.data_editor(
                    df.style.apply(lambda row: szinezo_admin(row, adatok, w), axis=1),
                    hide_index=True,
                    use_container_width=True,
                    key=f"editor_admin_{w}"
                )
                
                if st.button(f"Mentés: {w} szélesség", key=f"btn_save_{w}"):
                    for _, row in edited_df.iterrows():
                        kem = str(row.iloc[0])
                        if kem == "ÖSSZESEN":
                            continue
                        
                        for col in edited_df.columns[1:]:
                            col_str = str(col)
                            if col_str == "ÖSSZESEN":
                                continue
                                
                            val = row[col]
                            try:
                                new_val = int(val)
                            except (ValueError, TypeError):
                                new_val = 0
                                
                            sku = f"{col_str}_{w}_{kem}"
                            db.collection("keszlet").document(sku).set({"mennyiseg": new_val}, merge=True)
                    st.success(f"{w} szélesség készlete sikeresen elmentve!")
                    st.rerun()
    else: 
        st.warning("Add meg a jelszót!")
